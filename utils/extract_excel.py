from openpyxl import load_workbook
import logging

logger = logging.getLogger(__name__)

def extract_text(file_path):
    try:
        logger.info(f"Extracting text from Excel file: {file_path}")
        wb = load_workbook(file_path)
        text = ""
        for sheet in wb.sheetnames:
            logger.info(f"Processing sheet: {sheet}")
            ws = wb[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        text += str(cell.value) + " "
                text += "\n"
        logger.info(f"Successfully extracted text from Excel file: {file_path}")
        return text
    except Exception as e:
        logger.error(f"Error extracting text from Excel file {file_path}: {str(e)}")
        raise 